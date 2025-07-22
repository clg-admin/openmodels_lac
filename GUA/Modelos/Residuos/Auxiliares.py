import pandas as pd

def EliminarRepetidos(lista):
    salida=list()
    for i in range(len(lista)):
        if lista[i] not in salida:
            salida.append(lista[i])
    return salida

### Eliminar archivo
def EliminarArchivo(direccion_nombre):
    import os
    try:
        os.remove(direccion_nombre)
    except:
        pass   
        
def EliminarArchivosEnDirectorio(directorio):
    listaArchivos=ListaArchivos(directorio)
    import os
    for i in range(len(listaArchivos)):
        os.remove(directorio+"/"+listaArchivos[i])

#### Seccion de lectura de datos
def LeerExcel( ubicacion ):   
    datos = pd.ExcelFile( ubicacion )
    return datos

def LeerHoja2( excel , nombre , delay_filas ):
    data = excel.parse( nombre , skiprows = delay_filas )
    return data

def LeerHeaders( hoja ):
    headers = list( hoja )
    return headers

def LeerCol( hoja , nombre ):
    Col = list( hoja[nombre] )
    return Col

def ListaHojas( excel ):
    listaHojas = excel.sheet_names
    return listaHojas

def ListaArchivos( Nombre_carpeta ):
    import os
    listaArchivos = os.listdir( Nombre_carpeta )
    return listaArchivos

def LeerVariablesFinancieras( nombre_libro , nombre_hoja ):
    libro = LeerExcel( nombre_libro )
    hoja = LeerHoja2( libro, nombre_hoja, 0 )
    encab = LeerHeaders( hoja )
    tasa_descuento = LeerCol( hoja , encab[0] ) [0]
    tipo_de_cambio = LeerCol( hoja , encab[1] ) [0]
    tasa_interes_prestamo_viejo = LeerCol( hoja , encab[0] ) [5]
    tasa_interes_prestamo_nuevo = LeerCol( hoja , encab[1] ) [5]
    porc_finan = LeerCol( hoja , encab[0] ) [8]
    anios_prestamo = LeerCol( hoja , encab[0] ) [11]
    return tasa_descuento , tipo_de_cambio , tasa_interes_prestamo_nuevo , tasa_interes_prestamo_viejo , porc_finan , anios_prestamo

def TransponerTabla(matriz):
    trans=list()
    for i in range(len(matriz[0])):
        aux=list()
        for j in range(len(matriz)):
            aux.append(matriz[j][i])
        trans.append(aux)
    return trans

def MultLista(lista2,factor):
    lista=CopiaLista(lista2)
    for i in range(len(lista)):
        lista[i]=lista[i]*factor
    return lista

def CopiaLista(lista):
    copia=list()
    for i in lista:
        copia.append(i)
    return copia

def DetectarVectorCero(lista): ### NO era necesariamente verdad, por lo que se le agregó el abs()
    suma=0
    for i in lista:
        suma=suma+abs(i)
    if suma==0:
        return 1
    else:
        return 0

#### Seccion de variables económicas

def Financiamiento(tasa,years,monto): # tasa es un numero entre 0 y 1, years: cantidad de años, monto a financiar
    import numpy_financial as npf
    saldo=list()
    saldo.append(monto)
    intereses=list()
    intereses.append(0)
    amortizacion=list()
    amortizacion.append(0)
    pago=npf.pmt(tasa,years,-monto)

    for i in range(1,years+1):
        intereses.append(saldo[i-1]*tasa)
        amortizacion.append(pago-intereses[i])
        saldo.append(saldo[i-1]-amortizacion[i])

    return saldo, intereses, amortizacion, pago 


#serie_pagos_anuales, saldo, intereses, amortizacion = Financiamiento(0.13,10,10000)
#print(serie_pagos_anuales, saldo, intereses, amortizacion)
    
def ValorPresente(tasa_descuento,valores):
    import numpy_financial as npf
    enteros=Enteros(valores)
    salida=list()
    for i in range(len(enteros)):
        salida.append(npf.pv(tasa_descuento,enteros[i],0,-valores[i]))
    return salida

def Acumulado(valores):
    salida=list()
    for i in range(len(valores)):
        if i==0:
            salida.append(valores[i])
        else:
            salida.append(salida[len(salida)-1]+valores[i])
    return salida

def Enteros(lista):
    salida=list()
    for i in range(len(lista)):
        salida.append(i)
    return salida

def EnterosPositivos(lista):
    salida=list()
    for i in range(1,len(lista)+1):
        salida.append(i)
    return salida

# valores=[0.23001, 0.382143524, 0.381299587, 0.380143508, 0.379242952, 0.378189028, 0.148078031, 0.147094521, 0.145838355,
#          0.144410779, 0.144479736]
# valPres=ValorPresente(0.131,valores)
# print(valPres)

# Factor de recuperacion de capital
def FRC(tasa, periodo):
    #salida=(tasa * ((1 + tasa) ** (periodo+1))) / (((1 + tasa) ** (periodo+1)) - 1)
    salida=(tasa * ((1 + tasa) ** (periodo))) / (((1 + tasa) ** (periodo)) - 1)
    return salida

def CAE(VAN, fcr):
    return VAN*fcr

#print(FRC(0.131,11))


#### Seccion de escritura de archivos (resultados graficas)

def EscribirFilaEnHoja(datos, nombre_excel, nombre_hoja):
    dicColumnasExcel = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 14: "N", 
                        15: "O", 16: "P", 17: "Q", 18: "R", 19: "S", 20: "T", 21: "U", 22: "V", 23: "W", 24: "X", 25: "Y", 26: "Z"}
    import openpyxl
    book = openpyxl.load_workbook(nombre_excel)
    sheet = book[nombre_hoja]
    row_count = sheet.max_row
    for i in range(len(datos)):
        sheet[dicColumnasExcel[i+1] + str(row_count+1)]= datos[i]
    book.save(nombre_excel)
    
def CurvasEnExcel(nombre_excel, nombre_hoja, ubi):
    import openpyxl
    from openpyxl.chart import ScatterChart, Reference, Series
    dicColumnasExcel = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 14: "N", 
                        15: "O", 16: "P", 17: "Q", 18: "R", 19: "S", 20: "T", 21: "U", 22: "V", 23: "W", 24: "X", 25: "Y", 26: "Z"}
    book=openpyxl.load_workbook(nombre_excel)
    sheet = book[nombre_hoja]
    row_count = sheet.max_row
    column_count = sheet.max_column
    c1=ScatterChart()
    c1.y_axis.title="Flujos de Caja Acumulados (Millones de colones)"
    c1.x_axis.title="Años"
    xvalues=Reference(sheet, min_col=1, min_row=2, max_col=1, max_row=row_count)
    for i in range(2,column_count+1):
        yvalues=Reference(sheet, min_col=i, min_row=2, max_col=i, max_row=row_count)
        ser=Series(yvalues, xvalues, title=sheet[dicColumnasExcel[i] + "1"].value)
        c1.series.append(ser)
    sheet.add_chart(c1,ubi)
    book.save(nombre_excel)
    
def CurvasEnExcel2(nombre_excel, nombre_hoja, ubi, delay_filas, delay_columnas, casos, anios):
    import openpyxl
    from openpyxl.chart import ScatterChart, Reference, Series
    dicColumnasExcel = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 14: "N", 
                        15: "O", 16: "P", 17: "Q", 18: "R", 19: "S", 20: "T", 21: "U", 22: "V", 23: "W", 24: "X", 25: "Y", 26: "Z"}
    book=openpyxl.load_workbook(nombre_excel)
    sheet = book[nombre_hoja]
    
    c1=ScatterChart()
    c1.y_axis.title="Flujos de Caja Acumulados (Millones de colones)"
    c1.x_axis.title="Años"
    xvalues=Reference(sheet, min_col=delay_columnas, min_row=delay_filas+1, max_col=delay_columnas, max_row=delay_filas+1+anios)
    for i in range(casos):
        yvalues=Reference(sheet, min_col=delay_columnas+i+1, min_row=delay_filas+1, max_col=delay_columnas+i+1, max_row=delay_filas+1+anios)
        ser=Series(yvalues, xvalues, title=sheet[dicColumnasExcel[delay_columnas+i+1] + str(delay_filas)].value)
        c1.series.append(ser)
    
    sheet.add_chart(c1,ubi)
    book.save(nombre_excel)
    
def GraficaBarras(nombre_excel, nombre_hoja, ubi, delay_filas, delay_columnas, casos, num_desglose):
    import openpyxl
    from openpyxl.chart import BarChart, Reference

    book=openpyxl.load_workbook(nombre_excel)
    sheet = book[nombre_hoja]
    
    chart3 = BarChart()
    chart3.type = "col"
    #chart3.style = 12
    chart3.grouping = "stacked"
    chart3.overlap = 100
    chart3.y_axis.title = 'Valor actual neto (Millones de colones)'
    data = Reference(sheet, min_col=delay_columnas, min_row=delay_filas+1, max_row=num_desglose+delay_filas, max_col=delay_columnas+casos)
    cats = Reference(sheet, min_col=delay_columnas+1,max_col=delay_columnas+casos, min_row=delay_filas, max_row=delay_filas) # max_row=num_desglose+delay_filas)
    chart3.add_data(data, titles_from_data=True, from_rows=True)
    chart3.set_categories(cats)
    chart3.shape = 4
    sheet.add_chart(chart3, ubi)
    
    book.save(nombre_excel)

def BorrarLista(nombre_excel, nombre_hoja):
    import openpyxl
    book = openpyxl.load_workbook(nombre_excel)
    sheet = book[nombre_hoja]
    while sheet.max_row > 1:
        sheet.delete_rows(2)
    
    book.save(nombre_excel)
    
def SiglasFuel(string):
    if string==" Electricidad":
        salida=" ELEC"
    elif string==" Gasolina":
        salida=" GSL"
    elif string==" Diésel":
        salida=" DSL"
    else:
        salida=string
    return salida
    
    

#data=[0,1,2.0,"tres",4.2]
#EscribirFilaEnHoja(data, "TablaAcumulativaCasos.xlsx", "Hoja1")

#### Seccion de creacion de graficas
### Graficas de curvas de abatimiento
import matplotlib.pyplot as plt
import matplotlib
import random
matplotlib.rcParams.update({'font.size':15})
from matplotlib.lines import Line2D

def bubbleSort(arr,labels):
    n = len(arr)
    copia_arr=arr
    for i in range(n):
        for j in range(0, n-i-1):
            if copia_arr[j] > copia_arr[j+1] :
                copia_arr[j], copia_arr[j+1] = copia_arr[j+1], copia_arr[j] 
                labels[j], labels[j+1] = labels[j+1], labels[j]
    return copia_arr,labels
    
def autolabel(rects,colores, ax, decimales_leyenda_barras, separacion):
    """Attach a text label above each bar in *rects*, displaying its height."""
    cuenta=0
    for rect in rects:
        height = round(rect.get_height(), decimales_leyenda_barras)
        if height>0:
            ax.annotate('{}'.format(height),color=colores[cuenta],
                        xy=(rect.get_x() + rect.get_width() / 2, height*(1+separacion)),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom', rotation=90)
        else:
            ax.annotate('{}'.format(height),color=colores[cuenta],
                        xy=(rect.get_x() + rect.get_width() / 2, height*(1+separacion)),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='top', rotation=90)
        cuenta=cuenta+1

def CrearCurvaAbatimiento( etiquetas_barras, altura_barras, ancho_barras, multiplox_data, multiploy_data, tamx_fig, tamy_fig, etiquetax_eje, etiquetay_eje, nombre_guardado, multiplo_ejey_min, multiplo_ejey_max, tam_fuente_ejex, rotacion_etiquetas_ejex, cantidad_max_columnas_leyenda, int_localizacion_leyenda, ubi_leyendax, ubi_leyenday, decimales_ejex, decimales_leyenda_barras, separacion_leyenda_barra, eti_barra, ticks_ejex ):
    plt.figure(figsize=(tamx_fig,tamy_fig))
    ax = plt.subplot(111)
    
    colores=["#"+''.join([random.choice('0123456789ABCDEF') for e in range(6)])
            for p in range(len(etiquetas_barras))]
    
    height = MultLista(altura_barras, multiploy_data)
    
    width = MultLista(ancho_barras, multiplox_data)
    
    y_pos=list()
    y_pos2=list()
    acum=0
    for i in range(len(width)):
        y_pos.append(acum+width[i]/2)
        acum=acum+width[i]
        y_pos2.append(round(acum,decimales_ejex))    

    rects = ax.bar(y_pos, height, width=width, color=colores)

    custom_lines = list()
    for f in range(len(etiquetas_barras)):
        custom_lines.append(Line2D([0], [0], color=colores[f], lw=4))
    
    plt.legend(custom_lines,etiquetas_barras,bbox_to_anchor=(ubi_leyendax, ubi_leyenday),loc=int_localizacion_leyenda, ncol=cantidad_max_columnas_leyenda, borderaxespad=0.5)
    plt.xlim([0,y_pos2[len(y_pos2)-1]]) ## Cambiar
    plt.ylim([multiplo_ejey_min*min(height),multiplo_ejey_max*max(height)])
    if ticks_ejex==True:
        plt.xticks(y_pos2,y_pos2,rotation=rotacion_etiquetas_ejex,fontsize=tam_fuente_ejex)
    if eti_barra==True:
        autolabel(rects,colores, ax, decimales_leyenda_barras, separacion_leyenda_barra)
    plt.ylabel(etiquetay_eje)
    plt.xlabel(etiquetax_eje)
    plt.savefig( nombre_guardado , bbox_inches = 'tight' )
    plt.show()


def bubbleSort4D(arr,labels,labels2,labels3):
    n = len(arr)
    copia_arr=arr
    for i in range(n):
        for j in range(0, n-i-1):
            if copia_arr[j] > copia_arr[j+1] :
                copia_arr[j], copia_arr[j+1] = copia_arr[j+1], copia_arr[j] 
                labels[j], labels[j+1] = labels[j+1], labels[j]
                labels2[j], labels2[j+1] = labels2[j+1], labels2[j]
                labels3[j], labels3[j+1] = labels3[j+1], labels3[j]
    return(copia_arr,labels, labels2, labels3)

#### Seccion de lectura de datos
#### Lectura de datos de hojas especificas en dicionarios
# Conversion de litros a kwh de combustibles liquidos del sector aires y refrigeracion de sector transporte
def LecturaFactoresConversion_kWh_Litros(nombre_archivo_datos, nombre_hoja):
    libro=LeerExcel(nombre_archivo_datos)
    hoja=LeerHoja2(libro,nombre_hoja,1)
    encab=LeerHeaders(hoja)
    diesel=LeerCol(hoja, encab[0])[0]
    gasolina=LeerCol(hoja, encab[1])[0]
    elec=LeerCol(hoja, encab[2])[0]
    dicFactoresConversion_kWh_Litros={"Gasolina":gasolina,"Diésel":diesel, "Electricidad": elec}
    return dicFactoresConversion_kWh_Litros

# Leer Factores de Cambio de Refrigerante
def FactoresCambioRefrigerante(nombre_archivo_datos, nombre_hoja):
    libro=LeerExcel(nombre_archivo_datos)
    hoja=LeerHoja2(libro,nombre_hoja,0)
    encab=LeerHeaders(hoja)
    referencia=LeerCol(hoja, encab[0])
    opcion=LeerCol(hoja, encab[1])
    opcion_entre_referencia=LeerCol(hoja, encab[2])
    dicFactorCambioRefrigerante=dict()
    for i in range(len(referencia)):
        dicFactorCambioRefrigerante.update({referencia[i]+";"+opcion[i]:opcion_entre_referencia[i]})
    return dicFactorCambioRefrigerante

# Leer Relacion de Eficiencia Energetica EER en BTU/kWh
def EER_BTU_kWh(nombre_archivo_datos, nombre_hoja):
    libro=LeerExcel(nombre_archivo_datos)
    hoja=LeerHoja2(libro,nombre_hoja,0)
    encab=LeerHeaders(hoja)
    sec=LeerCol(hoja, encab[0])
    tec=LeerCol(hoja, encab[1])
    esce=LeerCol(hoja, encab[2])
    resto=encab[4:]
    dicEER=dict()
    for i in range(len(resto)):
        col=LeerCol(hoja,resto[i])
        for j in range(len(col)):
            dicEER.update({sec[j]+";"+tec[j]+";"+esce[j]+";"+str(resto[i]):col[j]})
    anios=resto
    return dicEER,anios


# Leer Precios de la Energia
def ObtenerPreciosEnergia(nombre_archivo_datos, nombre_hoja):
    libro=LeerExcel(nombre_archivo_datos)
    hoja=LeerHoja2(libro,nombre_hoja,1)
    encab=LeerHeaders(hoja)
    anio=LeerCol(hoja, encab[0])
    elec=LeerCol(hoja, encab[1])
    dsl=LeerCol(hoja, encab[2])
    gsl=LeerCol(hoja, encab[3])
    dicPreciosEnergia={encab[1]:elec,encab[2]:dsl,encab[3]:gsl}
    return dicPreciosEnergia, anio

# Leer Tabla de Opciones
def TablaOpciones(nombre_archivo_datos, nombre_hoja):
    libro=LeerExcel(nombre_archivo_datos)
    hoja=LeerHoja2(libro,nombre_hoja,0)
    encab=LeerHeaders(hoja)
    sec=LeerCol(hoja, encab[0])
    tec=LeerCol(hoja, encab[1])
    ref=LeerCol(hoja, encab[2])
    comb=LeerCol(hoja, encab[9])
    llaves=list()
    for i in range(len(sec)):
        llaves.append(sec[i]+";"+tec[i]+";"+ref[i]+";"+comb[i])

    dicOpcionesVidaUtil=dict(zip(llaves,LeerCol(hoja, encab[4])))
    dicOpcionesPCG=dict(zip(llaves,LeerCol(hoja, encab[5])))
    dicOpcionesCAPEX_dollar_TR=dict(zip(llaves,LeerCol(hoja, encab[6])))
    dicOpcionesFixed_prorcent_capex=dict(zip(llaves,LeerCol(hoja, encab[7])))
    return dicOpcionesVidaUtil , dicOpcionesPCG , dicOpcionesCAPEX_dollar_TR , dicOpcionesFixed_prorcent_capex


#### Lectura de datos de entrada
def LecturaDatosEntrada(nombre_archivo_datos_entrada):
    libro=LeerExcel(nombre_archivo_datos_entrada)
    nombre_hojas=ListaHojas(libro)
    hoja=LeerHoja2(libro,nombre_hojas[1],0) ### AQUI HAY QUE CAMBIAR
    encab=LeerHeaders(hoja)
    numero_caso=LeerCol(hoja, encab[0])
    sec=LeerCol(hoja, encab[1])
    tec=LeerCol(hoja, encab[2])
    ref=LeerCol(hoja, encab[3])
    comb=LeerCol(hoja, encab[4])    
    capacidad=LeerCol(hoja, encab[5])
    demanda=LeerCol(hoja, encab[6])
    carga=LeerCol(hoja, encab[7])
    llaves=list()
    for i in range(len(numero_caso)):
        llaves.append(str(numero_caso[i])+";"+sec[i]+";"+tec[i]+";"+ref[i]+";"+comb[i])
    dicDatosEntradaCapacidad=dict(zip(llaves,capacidad))
    dicDatosEntradaDemandaAnual=dict(zip(llaves,demanda))
    dicDatosEntradaCargaRefrigerante=dict(zip(llaves,carga))
    return dicDatosEntradaCapacidad,dicDatosEntradaDemandaAnual,dicDatosEntradaCargaRefrigerante


def MinimoMod(lista):
    lista2=CopiaLista(lista)
    for i in reversed(range(len(lista2))):
        if lista2[i] =="-":
            lista2.pop(i)
    return min(lista2)

def MaximoMod(lista):
    lista2=CopiaLista(lista)
    for i in reversed(range(len(lista2))):
        if lista2[i] =="-":
            lista2.pop(i)
    return max(lista2)
            
def ContadorEspecifico(lista,buscar):
    cuenta=0
    for i in range(len(lista)):
        if lista[i] == buscar:
            cuenta=cuenta+1
    return cuenta

def media(lista): # Calcula la media de una lista
  s = 0
  for elemento in lista:
    s += elemento
  return s / float(len(lista))